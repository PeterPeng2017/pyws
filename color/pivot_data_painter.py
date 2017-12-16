from compare.app_constants import AppConstants
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
import openpyxl


class PivotDataPainter:

    def __init__(self,
                 crm_file_path,
                 crm_sheet_name,
                 crm_result_file_path,
                 sap_file_path,
                 sap_sheet_name,
                 sap_result_file_path,
                 crm_data_container,
                 sap_data_container):

        self.crm_file_path = crm_file_path
        self.crm_sheet_name = crm_sheet_name
        self.crm_result_file_path = crm_result_file_path
        self.sap_file_path = sap_file_path
        self.sap_sheet_name = sap_sheet_name
        self.sap_result_file_path = sap_result_file_path
        self.crm_data_container = crm_data_container
        self.sap_data_container = sap_data_container

    def save(self):
        crm_file = openpyxl.load_workbook(self.crm_file_path)

        crm_product_sheet = crm_file.get_sheet_by_name(self.crm_sheet_name)

        self.paint_color(crm_product_sheet, self.crm_data_container)

        sap_file = openpyxl.load_workbook(self.sap_file_path)
        sap_product_sheet = sap_file.get_sheet_by_name(self.sap_sheet_name)
        self.paint_color(sap_product_sheet, self.sap_data_container)

        crm_file.save(self.crm_result_file_path)
        sap_file.save(self.sap_result_file_path)

    @staticmethod
    def paint_color(product_sheet, data_container):
        sheet_data = data_container.sheet_data
        for row_key, row_data in sheet_data.items():
            for cell_key, cell_unit in row_data.items():
                if cell_unit.color != AppConstants.NONE:
                    product_sheet.cell(row=cell_unit.x, column=cell_unit.y).fill \
                        = PatternFill("solid", fgColor=cell_unit.color)